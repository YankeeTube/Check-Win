from openpyxl.styles import Alignment, Font
from openpyxl.styles import PatternFill, Color
import openpyxl
import platform,codecs
import os,sys,re,time,datetime
import win32com.shell.shell as shell
sys.path.insert(0,'module')
import Mygetreg, servicepack, telnet_banner, hotfix, findvacine
import winstats,win32serviceutil
import requests
import pywintypes
import string
import wmi
from lxml import html

class Check_Vulnerability_Exposures:

    def __init__(self):
        self.result_list = {}
        self.status_code = {}

    def common_func(self,command):
        mycmd = os.popen("chcp 437 | {}".format(command))
        return mycmd
        # 2017.11.30(목) 15:53 - mycmd 뒤에 .read()제거 / 이유 : close()를 쓰기 위함 / 문제점 : close()를 안쓰면 프로세스가 안죽고 계속 자원을 갉아먹음

    def check_ShareFolder(self):
        csf = []
        alphabet = [alpha + "$" for alpha in string.ascii_uppercase[:26]]
        alphabet.append("IPC$")
        net = wmi.WMI()
        for share in net.Win32_share():
            if str(share.Name) not in alphabet:
                csf.append(share.Name)
        return csf

    def telnet_security(self):
        a = self.common_func("tlntadmn | find /i \"authentication mechanism\"")
        try:
            b = re.search(re.compile("(:\s+)(.*)"), a).group(2)
            if "NTLM" in b:
                b = b.replace("NTLM", "")
            result = b
        except:
            result =""
        return result

    def file_search(self):
        f = open("exception/[w-53]exception_list.txt", "r")
        f = f.read()
        filenames = os.listdir("C:/Windows/System32/Tasks/")
        location = "C:/Windows/System32/Tasks/"
        tasks = ""
        for filename in filenames:
            if filename in f:
                continue
            elif os.path.isdir(location + filename) is True:
                continue
            else:
                tasks = tasks + filename + "\n"
        return tasks

    def startUp_files(self):
        try:
            f = open("exception/[w-81]startup_list.txt", "r")
            f = f.read()
        except:
            f = ""
        startup_list = Mygetreg.startup_program()
        update_list = []
        for x in startup_list:
            if x in f:
                continue
            else:
                update_list.append(x)
        return update_list

    def startup_D(self,username):
        path = "C:/Users/{}/AppData/Roaming/Microsoft/Windows/Start Menu/Programs/Startup".format(username)
        try:
            f = open("exception/[w-81]startup_list.txt","r")
            f = f.read()
        except:
            f = ""
        update_list = []
        startup_folder = os.path.isdir(path)
        if startup_folder is True:
            filenames = os.listdir(path)
            for filename in filenames:
                if filename in f:
                    continue
                else:
                    update_list.append(filename)
        return update_list

    def program_list(self):
        program_name = command_line("wmic product get name")
        pn = program_name.split("\n")
        flag = False
        for line in pn:
            if "sql" in line or "SQL" in line:
                print(line)
                print("Success")
                flag = True
                break
        return flag

    # Administartor 계정 이름 바꾸기
    def w01(self):
        res = self.common_func(command="net localgroup administrators")
        try:
            result = re.search(re.compile("Administrator"), res.read()).group()
            result = "[W-01] Administrator 계정 이름 바꾸기 : " + str(result) + " 계정 존재"
            self.result_list['W-01'] = result
            self.status_code['W-01'] = "취약"
            res.close()
        except AttributeError:
            result = "[W-01] Administrator 계정 이름 바꾸기 : AdmInistrator 계정이 존재하지 않습니다."
            self.result_list['W-01'] = result
            self.status_code['W-01'] = "양호"
            res.close()
        except:
            result = "[W-01] Administrator 계정 이름 바꾸기 : None"
            self.result_list['W-01'] = result
            self.status_code['W-01'] = "양호"
            res.close()
        return result

    # Guest 계정 활성 상태
    def w02(self):
        res = self.common_func("net user guest")
        try:
            result = re.search(re.compile("active(.*)"),res.read()).group(1)
            result = result.replace(" ", "")
            result = result.replace("active","")
            if "No" in result:
                self.status_code['W-02'] = "양호"
                result = "[W-02] Guest 계정 상태 : " + str(result) + " [사용 안 함]"
            elif "Yes" in result:
                self.status_code['W-02'] = "취약"
                result = "[W-02] Guest 계정 상태 : " + str(result) + " [사용]"
            else:
                self.status_code['W-02'] = "비 정상"
                result = "[W-02] Guest 계정 상태 : 알 수 없음"
            self.result_list['W-02'] = result
        except AttributeError:
            result = "[W-02] Guest 계정 상태 : Guest 계정 사용 안함"
            self.result_list['W-02'] = result
            self.status_code['W-02'] = "양호"
        except:
            result = "[W-02] Guest 계정 상태 : None"
            self.result_list['W-02'] = result
            self.status_code['W-02'] = "양호"
        return result

    # 불필요한 계정 제거
    def w03(self):
        res = self.common_func('net user | find /v /i "User accounts" | find /v /i "command completed"')
        try:
            result = re.search(re.compile("A.*", re.DOTALL), res.read()).group().strip().replace("\n","")
            result = ", ".join(result.split())
            result = "[W-03] 불필요한 계정 : " + result
            self.result_list['W-03'] = result
            self.status_code['W-03'] = "취약"
        except AttributeError:
            result = "[W-03] 불필요한 계정 : 계정이 존재하지 않습니다."
            self.status_code['W-03'] = "양호"
            self.result_list['W-03'] = result
        except:
            result = "[W-03] 불필요한 계정 : None"
            self.result_list['W-03'] = result
            self.status_code['W-03'] = "양호"

    # 계정 잠금 입계값
    def w04(self):
        res = self.common_func("net accounts | find \"Lockout threshold\"")
        try:
            result = re.search(re.compile("L.*",re.DOTALL), res.read()).group()
            try:
                if re.search(re.compile(".*Never"),result).group() in result:
                    result = "[W-04] 계정 잠금 임계값 : Never [계정 미 잠금]"
                    self.result_list['W-04'] = result
                    self.status_code['W-04'] = "취약"
            except:
                result = re.search(re.compile("(?<=)\s[0-9.]+"), result).group()
                if int(result) < 5:
                    self.status_code['W-04'] = "취약"
                else:
                    self.status_code['W-04'] = "양호"
                result = "[W-04] 계정 잠금 임계값 : " + str(result) + "시도 가능"
                self.result_list['W-04'] = result
        except AttributeError:
            result = "[W-04] 계정 잠금 임계값 : 설정 값 없음 [계정 미 잠금]"
            self.status_code['W-04'] = "취약"
            self.result_list['W-04'] = result
        except:
            result = "[W-04] 계정 잠금 임계값 : 알 수 없음"
            self.result_list['W-04'] = result
            self.status_code['W-04'] = "비 정상"

    # 해독 가능한 암호화를 사용하여 암호 저장 [Policy 생성한 함수]
    def w05(self):
        try:
            res = self.common_func("secedit /EXPORT /CFG Local_Security_Policy.txt")  # 현재위치에 로컬 보안정책의 값을 텍스트파일로 생성
            res.close()
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("ClearTextPassword = ([0-9])"), result).group(1)
            if result == "0":
                self.status_code['W-05'] = "양호"
                result = "[W-05] 해독 가능한 암호화를 사용하여 암호 저장 : 사용"
            else:
                self.status_code['W-05'] = "취약"
                result = "[W-05] 해독 가능한 암호화를 사용하여 암호 저장 : 사용 안 함"
            self.result_list['W-05'] = result
        except AttributeError:
            result = "[W-05] 해독 가능한 암호화를 사용하여 암호 저장 : 사용 안 함"
            self.status_code['W-05'] = "취약"
            self.result_list['W-05'] = result
        except:
            result = "[W-05] 해독 가능한 암호화를 사용하여 암호 저장 : 알 수 없음"
            self.status_code['W-05'] = "비 정상"
            self.result_list['W-05'] = result

    # 관리자 그룹에 최소한의 사용자 포함
    def w06(self):
        res = self.common_func("net localgroup administrators")
        try:
            result = re.search(re.compile("-*(Administrator)(.*)(The command)", re.DOTALL),res.read()).group(2).replace("\n","")
            result = "[W-06] 관리자 그룹에 최소한의 사용자 포함 : " + result
            self.result_list['W-06'] = result
            self.status_code['W-06'] = "양호"
        except AttributeError:
            res.close()
            result = "[W-06] 관리자 그룹에 최소한의 사용자 포함 : 없음"
            self.result_list['W-06'] = result
            self.status_code['W-06'] = "취약"
        except:
            res.close()
            result = "[W-06] 관리자 그룹에 최소한의 사용자 포함 : 알 수 없음"
            self.result_list['W-06'] = result
            self.status_code['W-06'] = "비 정상"

    # Everyone 사용 권한을 익명 사용자에게 적용
    def w07(self):
        try:
            result = Mygetreg.w07_registry()
            if result == 0:
                self.status_code['W-07'] = "양호"
                result = "[W-07] Everyone 사용 권한을 익명 사용자에게 적용 :  [사용 안 함]"
            elif result == 1:
                self.status_code['W-07'] = "취약"
                result = "[W-07] Everyone 사용 권한을 익명 사용자에게 적용 :  [사용]"
            else:
                self.status_code['W-07'] = "비 정상"
                result = "[W-07] Everyone 사용 권한을 익명 사용자에게 적용 :  [알 수 없음]"
            self.result_list['W-07'] = result
        except:
            result = "[W-07] Everyone 사용 권한을 익명 사용자에게 적용 : 알 수 없음"
            self.result_list['W-07'] = result
            self.status_code['W-07'] = "취약"

    # 계정 잠금 기간 설정
    def w08(self):
        res1 = self.common_func("net accounts | find \"Lockout duration\"")
        res2 = self.common_func("net accounts | find \"observation\"")
        try:
            result1 = re.search(re.compile("([0-9]+)", re.DOTALL),res1.read()).group(1)
            result2 = re.search(re.compile("([0-9]+)",re.DOTALL),res2.read()).group(1)
            if int(result1) < 60 and int(result2) < 60:
                self.status_code['W-08'] = "취약"
                result = "[W-08] 계정 잠금 기간 설정\n계정 잠금 기간 [분] : " + str(result1) + "분 잠금 설정 사용 중\n계정 잠금 기간 원래대로 설정 기간 [분] : " + str(result2) + "분 잠금 설정 사용 중"
            else:
                self.status_code['W-08'] = "양호"
                result = "[W-08] 계정 잠금 기간 설정\n계정 잠금 기간 [분] : " + str(result1) + "분 잠금 설정 사용 중\n계정 잠금 기간 원래대로 설정 기간 [분] : " + str(result2) + "분 잠금 설정 사용 중"

            self.result_list['W-08'] = result
            res1.close()
            res2.close()
        except AttributeError:
            result = "[W-08] 계정 잠금 기간 설정[분] : 없음"
            self.result_list['W-08'] = result
            self.status_code['W-08'] = "취약"
            res1.close()
            res2.close()

        except:
            result = "[W-08] 계정 잠금 기간 설정[분] : Never [관리자의 해제 전까지 잠금]"
            self.result_list['W-08'] = result
            self.status_code['W-08'] = "양호"
            res1.close()
            res2.close()

    # 패스워드 복잡성 설정
    def w09(self):
        try:
            with open("Local_Security_Policy.txt", "r", encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("PasswordComplexity = ([0-9])"), result).group(1)
            if result == "0":
                self.status_code['W-09'] = "취약"
                result = "[W-09] 패스워드 복잡성 설정 : [사용 안 함]"
            else:
                self.status_code['W-09'] = "양호"
                result = "[W-09] 패스워드 복잡성 설정 : [사용]"
            self.result_list['W-09'] = result
        except AttributeError:
            result = "[W-09] 패스워드 복잡성 설정 : 설정 값 없음"
            self.result_list['W-09'] = result
            self.status_code['W-09'] = "취약"
        except:
            result = "[W-09] 패스워드 복잡성 설정 : 알 수 없음"
            self.result_list['W-09'] = result
            self.status_code['W-'] = "비 정 상"

    # 패스워드 최소 암호 길이
    def w10(self):
        try:
            with open("Local_Security_Policy.txt", "r", encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("MinimumPasswordLength = ([0-9])"), result).group(1)
            if result < "8":
                self.status_code['W-10'] = "취약"
                result = "[W-10] 패스워드 최소 암호 길이 : " + result + " 자리 미만 사용"
            else:
                self.status_code['W-10'] = "양호"
                result = "[W-10] 패스워드 최소 암호 길이 : " + result + " 자리 이상 사용중"
            self.result_list['W-10'] = result
        except AttributeError:
            result = "[W-10] 패스워드 최소 암호 길이 : 설정 값 없음"
            self.result_list['W-10'] = result
            self.status_code['W-10'] = "취약"
        except:
            result = "[W-10] 패스워드 최소 암호 길이 : 알 수 없음"
            self.result_list['W-10'] = result
            self.status_code['W-10'] = "비 정 상"

    # 패스워드 최대 사용 기간
    def w11(self):
        try:
            with open("Local_Security_Policy.txt", "r", encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("MaximumPasswordAge = ([0-9]+)"), result).group(1)
            if result < "90":
                self.status_code['W-11'] = "양호"
                result = "[W-11] 패스워드 최대 사용 기간 : " + result + "일 [권장사항 만족]"
            else:
                self.status_code['W-11'] = "취약"
                result = "[W-11] 패스워드 최대 사용 기간 : " + result + "일 [권장사항 불 만족]"
            self.result_list['W-11'] = result
        except AttributeError:
            result = "[W-11] 패스워드 최대 사용 기간 : 없음"
            self.result_list['W-11'] = result
            self.status_code['W-11'] = "취약"
        except:
            result = "[W-11] 패스워드 최대 사용 기간 : 알 수 없음"
            self.result_list['W-11'] = result
            self.status_code['W-11'] = "비 정 상"

    # 패스워드 최소 사용 기간
    def w12(self):
        try:
            with open("Local_Security_Policy.txt", "r", encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("MinimumPasswordAge = ([0-9]+)"), result).group(1)
            if result > "1":
                self.status_code['W-12'] = "양호"
                result = "[W-12] 패스워드 최소 사용 기간 : " + str(result) + "일 [권장사항 만족]"
            else:
                self.status_code['W-12'] = "취약"
                result = "[W-12] 패스워드 최소 사용 기간 : " + str(result) + "일 [권장사항 불 만족]"
            self.result_list['W-12'] = result
        except AttributeError:
            result = "[W-12] 패스워드 최소 사용 기간 : 설정 값 없음"
            self.result_list['W-12'] = result
            self.status_code['W-12'] = "취약"
        except:
            result = "[W-12] 패스워드 최소 사용 기간 : 알 수 없음"
            self.result_list['W-12'] = result
            self.status_code['W-12'] = "비 정 상"

    # 마지막 사용자 이름 표시 안 함
    def w13(self):
        try:
            result = Mygetreg.w13_registry()
            if result is not None:
                if result == 1:
                    self.status_code['W-13'] = "양호"
                    result = "[W-13] 마지막 사용자 이름 표시 안 함 : " + str(result) + " [사용]"
                elif result == 0:
                    self.status_code['W-13'] = "취약"
                    result = "[W-13] 마지막 사용자 이름 표시 안 함 : " + str(result) + " [사용 안 함]"
            else:
                self.status_code['W-13'] = "비 정 상"
                result = "[W-13] 마지막 사용자 이름 표시 안 함 : " + str(result) + " [확인 실패]"
            self.result_list['W-13'] = result

        except:
            result = "[W-13] 마지막 사용자 이름 표시 안 함 : 알 수 없음"
            self.result_list['W-13'] = result
            self.status_code['W-13'] = "비 정 상"

    # 로컬 로그온 허용
    def w14(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("SeInteractiveLogonRight = (.*)"),result).group(1).replace(",","\n")
            if "*S-1-5-32-544" in result:
                result = result.replace("*S-1-5-32-544","")
            if "*S-1-5-21" in result:
                result = result.replace("*S-1-5-21","")
            if "IUSR_" in result:
                result = result.replace("IUSR_","")
            result = result.replace(",","")
            if len(result) > 0:
                self.status_code['W-14'] = "취약"
                result = "[W-14] 로컬 로그온 허용 : \n" + result.replace("\n\n","\n")
            else:
                self.status_code['W-'] = "양호"
                result = "[W-14] 로컬 로그온 허용 : None"
            self.result_list['W-14'] = result
        except AttributeError:
            result = "[W-14] 로컬 로그온 허용 : 허용 계정 없음"
            self.result_list['W-14'] = result
            self.status_code['W-14'] = "양호"
        except:
            result = "[W-14] 로컬 로그온 허용 : 알 수 없음"
            self.result_list['W-14'] = result
            self.status_code['W-14'] = "비 정 상"

    # 익명 SID/이름 변환 허용
    def w15(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("LSAAnonymousNameLookup = ([0-9]+)"),result).group(1)
            if int(result) > 1: # 허용
                self.status_code['W-15'] = "취약"
                result = "[W-15] 익명 SID/이름 변환 허용 : [허용]"
            else: # 허용안함
                self.status_code['W-15'] = "양호"
                result = "[W-15] 익명 SID/이름 변환 허용 : [허용 안 함]"
            self.result_list['W-15'] = result
        except AttributeError:
            result = "[W-15] 익명 SID/이름 변환 허용 : 설정 값 없음"
            self.result_list['W-15'] = result
            self.status_code['W-15'] = "양호"
        except:
            result = "[W-15] 익명 SID/이름 변환 허용 : 알 수 없음"
            self.result_list['W-15'] = result
            self.status_code['W-15'] = "비 정 상"

    # 최근 암호 기억
    def w16(self):
        res = self.common_func("net accounts | find \"history\"")
        try:
            result = re.search(re.compile("(:)(.*)", re.DOTALL),res.read()).group(2).replace(" ", "").replace("\n","")
            if "None" in result:
                self.status_code['W-16'] = "취약"
                result = "[W-16] 최근 암호 기억 : 설정 값 없음"
            elif int(result) <= 12:
                self.status_code['W-16'] = "취약"
                result = "[W-16] 최근 암호 기억 : " + result + "일"
            else:
                self.status_code['W-16'] = "양호"
                result = "[W-16] 최근 암호 기억 : " + result + "일"
            self.result_list['W-16'] = result
            res.close()
        except AttributeError:
            result = "[W-16] 최근 암호 기억 : 설정 값 없음"
            self.result_list['W-16'] = result
            self.status_code['W-16'] = "취약"
            res.close()
        except:
            result = "[W-16] 최근 암호 기억 : 확인 실패"
            self.result_list['W-16'] = result
            self.status_code['W-16'] = "비 정 상"
            res.close()

    # 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한
    def w17(self):
        #0 사용안함
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("LimitBlankPasswordUse=4,([0-9])"),result).group(1)
            if int(result) == 1:
                self.status_code['W-17'] = "양호"
                result = "[W-17] 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 : [사용]"
            else:
                self.status_code['W-17'] = "취약"
                result = "[W-17] 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 : [사용 안 함]"
            self.result_list['W-17'] = result
        except AttributeError:
            result = "[W-17] 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 : 설정 값 없음"
            self.result_list['W-17'] = result
            self.status_code['W-17'] = "취약"
        except:
            result = "[W-17] 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 : 알 수 없음"
            self.result_list['W-17'] = result
            self.status_code['W-17'] = "비 정 상"

    # 원격터미널 접속 가능한 사용자 그룹 제한
    def w18(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("SeRemoteInteractiveLogonRight = (.*)"),result).group(1)
            if "*S-1-5-32-544" in result:
                result = result.replace("*S-1-5-32-544","")
            if "*S-1-5-32-555" in result:
                result = result.replace("*S-1-5-32-555","")
            result = result.replace(",","")
            if len(result) < 1:
                self.status_code['W-18'] = "취약"
                result = "[W-18] 원격터미널 접속 가능한 사용자 그룹 제한 : 제한 등록된 사용자 및 그룹이 존재하지 않습니다."
            else:
                self.status_code['W-18'] = "양호"
                result = "[W-18] 원격터미널 접속 가능한 사용자 그룹 제한 : " + str(result) + "등록됨"
            self.result_list['W-18'] = result
        except AttributeError:
            result = "[W-18] 원격터미널 접속 가능한 사용자 그룹 제한 : 설정 값 없음"
            self.result_list['W-18'] = result
            self.status_code['W-18'] = "취약"
        except:
            result = "[W-18] 원격터미널 접속 가능한 사용자 그룹 제한 : 알 수 없음"
            self.result_list['W-18'] = result
            self.status_code['W-18'] = "비 정 상"

# ---------------------- 2. 서비스 관리 -------------------------------#

    # 공유 권한 및 사용자 그룹 설정
    def w19(self):
        res = self.common_func("net share  | FIND /v \"$\" | FIND /i \"c:\\\"")
        try:
            if len(res.read()) > 0:
                self.status_code['W-19'] = "취약"
                self.common_func("powershell Set-ExecutionPolicy bypass")
                name = self.common_func("powershell.exe -NoProfile -ExecutionPolicy RemoteSigned -File module/auth.ps1 | find \"Share name\"").read().split("\n")
                auth = self.common_func("powershell.exe -NoProfile -ExecutionPolicy RemoteSigned -File module/auth.ps1 | find \"Permission\"").read().split("\n")
                share_auth = ""
                try:
                    del name[-1]
                    del auth[-1]
                    for x in range(len(name)):
                        try:
                            if "$" in name[x]:
                                del name[x]
                                del auth[x]
                            else:
                                share_auth += "-" * 50 + "\n" + name[x] + "\n" + auth[x] + "\n" + "-" * 25
                        except IndexError:
                            break
                    for y in range(len(name)):
                        share_auth += "-" * 50 + "\n" + name[y] + "\n" + auth[y] + "\n" + "-" * 25
                except:
                    name.close()
                    auth.close()
                    pass
                result = "[W-19] 공유 권한 및 사용자 그룹 설정 : \n" + share_auth
                self.result_list['W-19'] = result
            else:
                result = "[W-19] 공유 권한 및 사용자 그룹 설정 : 없음"
                self.result_list['W-19'] = result
                self.status_code['W-19'] = "양호"
                res.close()
        except:
            result = "[W-19] 공유 권한 및 사용자 그룹 설정 : 알 수 없음"
            self.result_list['W-19'] = result
            self.status_code['W-19'] = "비 정 상"
            res.close()
            pass

    # 하드디스크 기본 공유 제거
    def w20(self):
        try:
            result1 = Mygetreg.w20_registry()
            result2 = self.check_ShareFolder()
            if str(result1) == "0" or len(result1) < 1 and len(result2) < 1:
                self.status_code['W-20'] = "양호"
                result = "[W-20] 하드디스크 기본 공유 제거 : 기본 공유폴더를 사용하고 있지 않습니다."
            else:
                self.status_code['W-20'] = "취약"
                result = "[W-20] 하드디스크 기본 공유 제거 : 레지스트리 설정이 없거나 기본 공유폴더를 사용중입니다."
            self.result_list['W-20'] = result
        except:
            self.result_list['W-20'] = result
            self.status_code['W-20'] = "비 정 상"
            result = "[W-20] 하드디스크 기본 공유 제거 : 알 수 없음"

    # 불 필요한 서비스 제거
    def w21(self):
        res1 = self.common_func("net start | findstr \"Alerter ClipBook Messenger\"")
        res2 = self.common_func("net start | find \"Simple TCP/IP Services\"")
        try:
            if len(res1.read()) > 1:
                if len(res2.read()) > 1:
                    self.status_code['W-21'] = "취약"
                    result = "[W-21] 불 필요한 서비스 제거 : " + str(res1.read()) + " " + str(res2.read())
                    self.result_list['W-21'] = result
            else:
                self.status_code['W-21'] = "양호"
                result = "[W-21] 불 필요한 서비스 제거 : 없음"
                self.result_list['W-21'] = result
            res1.close()
            res2.close()

        except:
            result = "[W-21] 불 필요한 서비스 제거 : 알 수 없음"
            self.result_list['W-21'] = result
            self.status_code['W-21'] = "비 정 상"
            res1.close()
            res2.close()

    # IIS 서비스 구동 점검
    def w22(self):
        res = self.common_func("dir %systemroot%\System32\inetsrv\config\schema")
        try:
            if "지정된 파일을 찾을 수 없습니다." in res.read():
                self.status_code['W-22'] = "양호"
                result = "[W-22] IIS 서비스 구동 점검 : " + str(res.read())
            elif "" in res.read():
                self.status_code['W-22'] = "양호"
                result = "[W-22] IIS 서비스 구동 점검 : IIS 서버 존재하지 않음."
            else:
                self.status_code['W-22'] = "취약"
                result = "[W-22] IIS 서비스 구동 점검 : \n" + re.search(re.compile("Directory(.*)schema"),str(res.read())).group(1)
            self.result_list['W-22'] = result
            res.close()
        except:
            result = "[W-22] IIS 서비스 구동 점검 : 알 수 없음"
            self.result_list['W-22'] = result
            self.status_code['W-22'] = "비 정 상"
            res.close()

    # IIS 디렉토리 리스팅 제거
    def w23(self):
        res = self.common_func("type %systemroot%\system32\inetsrv\config\\applicationHost.config | find /I \"directoryBrowse enabled\"")
        try:
            result = re.search(re.compile("(\")(.*)(\")"), res.read()).group(2)
            if "false" in res or "False" in res or "FALSE" in res:
                self.status_code['W-23'] = "양호"
                result = "[W-23] IIS 디렉토리 리스팅 제거 : " + str(result)
            else:
                self.status_code['W-23'] = "취약"
                result = "[W-23] IIS 디렉토리 리스팅 제거 : " + str(result)
            self.result_list['W-23'] = result
            res.close()
        except AttributeError:
            result = "[W-23] IIS 디렉토리 리스팅 제거 : IIS 사용 안 함"
            self.result_list['W-23'] = result
            self.status_code['W-23'] = "양호"
            res.close()
        except:
            result = "[W-23] IIS 디렉토리 리스팅 제거 : 알 수 없음"
            self.result_list['W-23'] = result
            self.status_code['W-23'] = "비 정 상"
            res.close()

    # IIS CGI 실행 제한
    def w24(self):
        res = self.common_func("icacls c:\inetpub\scripts | find /i \"everyone\"")
        try:
            result = os.path.exists("c:\inetpub\scripts")
            if result == False:
                self.status_code['W-24'] = "양호"
                result = "[W-24] IIS CGI 실행 제한 : " + str(result)
            else:
                if len(res.read()) > 1:
                    self.status_code['W-24'] = "취약"
                    result = "[W-24] IIS CGI 실행 제한 : " + str(res.read())
                else:
                    self.status_code['W-24'] = "양호"
                    result = "[W-24] IIS CGI 실행 제한 : 제한 안 함"
            self.result_list['W-24'] = result
            res.close()
        except:
            result = "[W-24] IIS CGI 실행 제한 : 알 수 없음"
            self.result_list['W-24'] = result
            self.status_code['W-24'] = "비 정 상"
            res.close()

    # IIS 상위 디렉터리 접근 금지
    def w25(self):
        try:
            with open("{}/WINDOWS/System32/inetsrv/config/applicationHost.config".format(os.getenv("SystemDrive")),"r",encoding="utf-8") as f:
                res = f.read()
                try:
                    res = re.search(re.compile("enableParentPaths=(.*)"),res).group(1)
                except:
                    self.status_code['W-25'] = "양호"
                    result = "[W-25] IIS 상위 디렉터리 접근 금지 : 사용"
            try:
                if len(res) > 1:
                    self.status_code['W-25'] = "취약"
                    result = "[W-25] IIS 상위 디렉터리 접근 금지 : " + str(res)
                elif "" in res:
                    self.status_code['W-25'] = "양호"
                    result = "[W-25] IIS 상위 디렉터리 접근 금지 : IIS 서버 사용 안함."
                else:
                    self.status_code['W-25'] = "양호"
                    result = "[W-25] IIS 상위 디렉터리 접근 금지 : 사용"
                self.result_list['W-25'] = result
            except:
                result = "[W-25] IIS 상위 디렉터리 접근 금지 : 사용"
                self.result_list['W-25'] = result
                self.status_code['W-25'] = "비 정 상"
        except:
            result = "[W-25] : IIS Server를 사용하고 있지 않습니다."
            self.result_list['W-25'] = result
            self.status_code['W-25'] = "비 정 상"

    #IIS 불필요한 파일 제거
    def w26(self):
        res = Mygetreg.iis_version()
        if res is not None:
            try:
                if float(res) > 6.0:
                    self.status_code['W-26'] = "양호"
                    result = "[W-26] IIS 불필요한 파일 제거 : IIS Version-" + str(res) + "은 Sample 파일이 존재하지 않습니다."
                else:
                    res1 = self.common_func("type c:\windows\system32\inetsrv\config\\applicationHost.config | findstr /l  IISSamples")
                    res2 = self.common_func("type c:\windows\system32\inetsrv\config\\applicationHost.config | findstr /l IISHelp")
                    res3 = self.common_func("findstr /l  IISSamples c:\windows\system32\inetsrv\config\applicationHost.config > nul")
                    res4 = self.common_func("findstr /l  IISSamples c:\windows\system32\inetsrv\config\applicationHost.config > nul")

                    if len(res1) > 1 or len(res2) > 1 or len(res3) > 1 or len(res4) > 1:
                        self.status_code['W-26'] = "취약"
                        result = "[W-26] IIS 불필요한 파일 제거 : " + str(res1) +"\n"+str(res2)+"\n"+str(res3)+"\n"+str(res4)
                    else:
                        self.status_code['W-26'] = "양호"
                        result = "[W-26] IIS 불필요한 파일 제거 : 없음"
                self.result_list['W-26'] = result
            except:
                result = "[W-26] IIS 불필요한 파일 제거 : IIS 서버 동작 안 함."
                self.status_code['W-26'] = "비 정 상"
                self.result_list['W-26'] = result
        else:
            result = "[W-26] IIS 불필요한 파일 제거 : IIS 서버 동작 안 함."
            self.status_code['W-26'] = "비 정 상"
            self.result_list['W-26'] = result

    # IIS 웹 프로세스 권한 제한
    def w27(self):
        try:
            result = Mygetreg.w27_registry()
            if result == "LocalSystem":
                result = "[W-27] IIS 웹 프로세스 권한 제한 : " + str(result)
                self.status_code['W-27'] = "취약"
            elif result == False:
                result = "[W-27] IIS 웹 프로세스 권한 제한 : 없음"
                self.status_code['W-27'] = "양호"
            else:
                result = "[W-27] IIS 웹 프로세스 권한 제한 : " + str(result)
                self.status_code['W-27'] = "양호"
            self.result_list['W-27'] = result
        except:
            result = "[W-27] IIS 웹 프로세스 권한 제한 : 알 수 없음"
            self.result_list['W-27'] = result
            self.status_code['W-27'] = "비 정 상"

    # IIS 링크 사용 금지 - 미완?
    def w28(self):
        res = self.common_func("type {}/windows/system32/inetsrv/config/applicationHost.config | findstr /i \"physicalPath\" | findstr /i \".lnk\"")
        try:
            if len(res.read()) > 1:
                result = "[W-28] IIS 링크 사용 금지 : 홈 디렉터리 바로가기 링크 사용중"
                self.status_code['W-28'] = "취약"
            else:
                result = "[W-28] IIS 링크 사용 금지 : 홈 디렉터리 바로가기 링크 사용 안 함"
                self.status_code['W-28'] = "양호"
            self.result_list['W-28'] = result
            res.close()
        except:
            result = "[W-28] IIS 링크 사용 금지 : 알 수 없음"
            self.result_list['W-28'] = result
            self.status_code['W-28'] = "비 정 상"
            res.close()

    # IIS 파일 업로드 및 다운로드 제한
    def w29(self):
        res1 = self.common_func("type c:\windows\system32\inetsrv\config\applicationHost.config | find /i \"bufferingLimit\"")
        try:
            if len(res1.read()) > 1:
                res2 = self.common_func("type c:\windows\system32\inetsrv\config\applicationHost.config | find /i \"maxRequestEntityAllowed\"")
                if len(res2.read()) > 1:
                    result = "[W-29] IIS 파일 업로드 및 다운로드 제한 : 업로드 / 다운로드 제한 적용됨"
                    self.status_code['W-29'] = "양호"
                else:
                    result = "[W-29] IIS 파일 업로드 및 다운로드 제한 : 업로드 제한 적용됨"
                    self.status_code['W-29'] = "취약"
                res2.close()
            else:
                result = "[W-29] IIS 파일 업로드 및 다운로드 제한 : 업로드 / 다운로드 제한 적용 안 함"
                self.status_code['W-29'] = "취약"
            self.result_list['W-29'] = result
            res1.close()
        except:
            result = "[W-29] IIS 파일 업로드 및 다운로드 제한 : 알 수 없음"
            self.result_list['W-29'] = result
            self.status_code['W-29'] = "비 정 상"
            res1.close()

    # IIS DB 연결 취약점 점검
    def w30(self):
        res = self.common_func("type %systemroot%\system32\inetsrv\config\\applicationHost.config | findstr /I .asa")
        try:
            result = re.search(re.compile("(allowed=\")(.*)(\")"), res.read()).group(2)
            if "false" in result or "False" in result or "FALSE" in result:
                self.status_code['W-30'] = "양호"
                result = "[W-30] IIS DB 연결 취약점 점검 : " + str(result)
            else:
                self.status_code['W-30'] = "취약"
                result = "[W-30] IIS DB 연결 취약점 점검 : " + str(result)
            self.result_list['W-30'] = result
            res.close()
        except AttributeError:
            result = "[W-30] IIS DB 연결 취약점 점검 : 설정 값 없음"
            self.result_list['W-30'] = result
            self.status_code['W-30'] = "취약"
            res.close()
        except:
            result = "[W-30] IIS DB 연결 취약점 점검 : 알 수 없음"
            self.result_list['W-30'] = result
            self.status_code['W-'] = "비 정 상"
            rse.close()

    # IIS 가상 디렉터리 삭제
    def w31(self):
        res1 = self.common_func("%systemroot%\system32\inetsrv\\appcmd list vdir")
        try:
            result = re.search(re.compile("(admin)(.*)"),res1.read()).group(1)
            if "admin" in result:
                res2 = self.common_func("%systemroot%\system32\inetsrv\\appcmd list vdir")
                try:
                    result2 = re.search(re.compile("(adminpwd)(.*)"),res2.read()).group(1)
                    self.status_code['W-31'] = "취약"
                    result = "[W-31] IIS 디렉터리 삭제 : \n" + str(result) + ", " + str(result) + " 존재"
                    res2.close()
                except:
                    self.status_code['W-'] = "취약"
                    result = "[W-31] IIS 디렉터리 삭제 : \n" + str(result) + " 존재"
                    res2.close()
            else:
                result = "[W-31] IIS 디렉터리 삭제 : 가상 디렉터리 없음"
                self.status_code['W-31'] = "양호"
            self.result_list['W-31'] = result
            res1.close()
        except:
            result = "[W-31] IIS 디렉터리 삭제 : 알 수 없음"
            self.result_list['W-31'] = result
            self.status_code['W-31'] = "비 정 상"
            res1.close()

    # IIS 데이터 파일 ACL 적용
    def w32(self):
        res = self.common_func("icacls C:\Inetpub\wwwroot | find /I \"everyone\"")
        try:
            if len(res.read()) > 1:
                result = "[W-32] IIS 데이터 파일 ACL 적용 : " + str(res.read())
                self.status_code['W-32'] = "취약"
            else:
                result = "[W-32] IIS 데이터 파일 ACL 적용 : Everyone 권한 파일이 존재 하지 않음"
                self.status_code['W-32'] = "양호"
            self.result_list['W-32'] = result
            res.close()
        except:
            result = "[W-32] IIS 데이터 파일 ACL 적용 : 알 수 없음"
            self.result_list['W-32'] = result
            self.status_code['W-32'] = "비 정 상"
            res.close()

    # IIS 미사용 스크립트 매핑 제거
    def w33(self):
        try:
            execute = [".htr",".idc","stm",".shtm",".printer",".htw",".ida",".idq"]
            result = []
            for x in execute:
                res = self.common_func("%systemroot%\system32\inetsrv\\appcmd list config | findstr /i \"{}\"".format(x))
                result.append(res.read())
                res.close()
            if len(result) > 1:
                result = "[W-33] IIS 미사용 스크립트 매핑 제거 : " + str(result) + " 발견"
                self.status_code['W-33'] = "취약"
            else:
                result = "[W-33] IIS 미사용 스크립트 매핑 제거 : " + str(result) + " 발견"
                self.status_code['W-33'] = "양호"
            self.result_list['W-33'] = result
        except:
            result =  "[W-33] IIS 미사용 스크립트 매핑 제거 : 알 수 없음"
            self.result_list['W-33'] = result
            self.status_code['W-33'] = "비 정 상"

    # IIS Exec 명령어 쉘 호출 진단
    def w34(self):
        try:
            result = Mygetreg.w34_registry()
            if result == 1:
                self.status_code['W-34'] = "취약"
                result = "[W-34] IIS Exec 명령어 쉘 호출 진단 : " + str(result) + " [사용]"
            elif result == 0:
                self.status_code['W-34'] = "양호"
                result = "[W-34] IIS Exec 명령어 쉘 호출 진단 : " + str(result) + " [사용 안 함]"
            else:
                self.status_code['W-34'] = "비 정 상"
                result = "[W-34] IIS Exec 명령어 쉘 호출 진단 : 알 수 없음"
            self.result_list['W-34'] = result
        except:
            result = "[W-34] IIS Exec 명령어 쉘 호출 진단 : 파일이 존재하지 않습니다."
            self.result_list['W-34'] = result
            self.status_code['W-34'] = "양호"

    # IIS WebDAV 비 활성화
    def w35(self):
        try:
            result = Mygetreg.w34_registry()
            if len(result) > 1:
                result = "[W-35] IIS WebDAV 비 활성화 : " + str(result) + "[사용]"
                self.status_code['W-35'] = "양호"
            else:
                result = "[W-35] IIS WebDAV 비 활성화 : " + str(result) + "[사용 안 함]"
                self.status_code['W-35'] = "취약"
            self.result_list['W-35'] = result
        except:
            version = Mygetreg.iis_version()
            result = "[W-35] IIS WebDAV 비 활성화 : IIS-Version : " + str(version) + " 은 점검 대상 버전이 아닙니다."
            self.result_list['W-35'] = result
            self.status_code['W-35'] = "비 정 상"

    # NetBIOS 바인딩 서비스 구동 점검
    def w36(self):
        result = Mygetreg.netBios_Bind_Service()
        if result is not None:
            try:
                if result == 0: # 기본값
                    result = "[W-36] NetBIOS 바인딩 서비스 구동 점검 : " + str(result) + "[기본값]"
                    self.status_code['W-36'] = "취약"
                elif result == 1: # 사용
                    result = "[W-36] NetBIOS 바인딩 서비스 구동 점검 : " + str(result) + "[사용]"
                    self.status_code['W-36'] = "취약"
                else: # 2 사용 안 함
                    result = "[W-36] NetBIOS 바인딩 서비스 구동 점검 : " + str(result) + "[사용 안 함]"
                    self.status_code['W-36'] = "양호"
                self.result_list['W-36'] = result
            except:
                result = "[W-36] NetBIOS 바인딩 서비스 구동 점검 : 알 수 없음"
                self.result_list['W-36'] = result
                self.status_code['W-36'] = "비 정 상"
        else:
            result = "[W-36] NetBIOS 바인딩 서비스 구동 점검 : 알 수 없음"
            self.result_list['W-36'] = result
            self.status_code['W-36'] = "비 정 상"

    # FTP 서비스 구동 점검
    def w37(self):
        res = self.common_func("net start | find /i \"FTP Service\"")
        try:
            if len(res.read()) > 1:
                result = "[W-37] FTP 서비스 구동 점검 : " + str(res.read())
                self.status_code['W-37'] = "취약"
            else:
                result = "[W-37] FTP 서비스 구동 점검 : " + str(res.read())
                self.status_code['W-37'] = "양호"
            self.result_list['W-37'] = result
            res.close()
        except:
            result = "[W-37] FTP 서비스 구동 점검 : 알 수 없음"
            self.result_list['W-37'] = result
            self.status_code['W-37'] = "비 정 상"
            res.close()

    # FTP 디렉토리 접근 권한 설정
    def w38(self):
        res = self.common_func("type C:\Windows\System32\inetsrv\config\\applicationHost.config | find /i \"physicalPath\"")
        try:
            result = re.search(re.compile("(physicalPath=)(.*)",re.DOTALL),res).group(2)
            result = "[W-38] FTP 디렉토리 접근 권한 설정 : " + result + "[사용중]"
            self.status_code['W-38'] = "취약"
            self.result_list['W-38'] = result
        except AttributeError:
            result = "[W-38] FTP 디렉토리 접근 권한 설정 : 설정 값 없음"
            self.result_list['W-38'] = result
            self.status_code['W-38'] = "양호"
        except:
            result = "[W-38] FTP 디렉토리 접근 권한 설정 : 알 수 없음"
            self.result_list['W-38'] = result
            self.status_code['W-38'] = "비 정 상"

    # Anonymouse FTP 금지
    def w39(self):
        res = self.common_func("type C:\Windows\System32\inetsrv\config\\applicationHost.config | find /i \"anonymousAuthentication enabled\"")
        try:
            result = re.search(re.compile("(enabled=)(.*)(\s)",re.DOTALL),res.read())
            result = "[W-39] Anonymous FTP 금지 : " + str(result)
            self.result_list['W-39'] = result
            self.status_code['W-39'] = "취약"
            res.close()
        except AttributeError:
            result = "[W-39] Anonymous FTP 금지 : 설정 값 없음"
            self.result_list['W-39'] = result
            self.status_code['W-39'] = "양호"
            res.close()
        except:
            result = "[W-39] Anonymous FTP 금지 : 알 수 없음"
            self.result_list['W-39'] = result
            self.status_code['W-39'] = "비 정 상"
            res.close()

    # FTP 접근 제어 설정
    def w40(self):
        res = self.common_func("type C:\Windows\System32\inetsrv\config\\applicationHost.config | find /i \"add ipAddress\"")
        try:
            result = re.search(re.compile("(add)(.*)(/)",re.DOTALL),res.read()).group(2)
            if len(result) > 1:
                result = "[W-40] FTP 접근 제어 설정 : " + str(result) + " />"
                self.status_code['W-40'] = "취약"
            else:
                result = "[W-40] FTP 접근 제어 설정 : 없음"
                self.status_code['W-40'] = "양호"
            self.result_list['W-40'] = result
            res.close()
        except:
            result = "[W-40] FTP 접근 제어 설정 : 알 수 없음"
            self.result_list['W-40'] = result
            self.status_code['W-40'] = "비 정 상"
            res.close()

    # DNS Zone Transfer 설정
    def w41(self):
        try:
            result = Mygetreg.w41_registry()
            if len(result) > 1:
                for domain in result.keys():
                    setSecurity = result.get(domain).get('SecureSecondaries')
                    if setSecurity == 1:
                        result = "[W-41] DNS Zone Transfer 설정 : 영역 전송 허용 안 함\n" + str(domain)
                        self.status_code['W-41'] = "양호"
                    elif setSecurity == 2:
                        server2 = result.get(domain).get('SecondaryServers')
                        result = "[W-41] DNS Zone Transfer 설정 : 특정 서버로만 전송\n" + str(domain) +"\n" + str(server2)
                    else:
                        result = "[W-41] DNS Zone Transfer 설정 : 모두 전송 허용"
                        self.status_code['W-41'] = "취약"
            else:
                result = "[W-41] DNS Zone Transfer 설정 : DNS 사용 안 함"
                self.status_code['W-41'] = "양호"
            self.result_list['W-41'] = result
        except:
            result = "[W-41] DNS Zone Transfer 설정 : DNS Zone이 존재하지 않습니다."
            self.result_list['W-41'] = result
            self.status_code['W-41'] = "양호"

    # RDS(RemoteDataServices)제거
    def w42(self):
        try:
            locate = "{}\Program Files\Common Files\System\msadc\msadcs.dll".format(os.getenv("SystemDrive"))
            locate = os.path.isfile(locate)
            locate2 = Mygetreg.w42_registry()
            if locate is False and locate2 is False:
                result = "[W-42] : RDS(RemoteDataServices)제거 : RDS 인터페이스가 존재하지 않음"
                self.status_code['W-42'] = "양호"
            else:
                result = "[W-42] : RDS(RemoteDataServices)제거 : RDS 인터페이스가 존재 함"
                self.status_code['W-42'] = "취약"
            self.result_list['W-42'] = result
        except:
            result = "[W-42] : RDS(RemoteDataServices)제거 : 알 수 없음"
            self.status_code['W-42'] = "비 정 상"
            self.result_list['W-42'] = result

    # 최신 서비스팩 적용
    def w43(self):
        result = servicepack.check_ServicePack()
        if type(result) == "str":
            res = "[W-43] 최신 서비스팩 적용 : \n" + result
            self.result_list['W-43'] = res
            self.status_code['W-43'] = "비 정 상"
        else:
            res = "[W-43] 최신 서비스팩 적용 : \n" + result[0]
            self.result_list['W-43'] = res
            self.status_code['W-43'] = result[1]

    # 터미널 서비스 암호화 수준 설정
    def w44(self):
        res = self.common_func("NET START | FIND \"Terminal Service\" > NUL")
        try:
            if len(res.read()) > 1:
                try:
                    value = Mygetreg.w44_registry() # RDP-TCP 최소 암호화 수준 확인 / False면 None
                    if value is not None:
                        if int(value) < 2:
                            result = "[W-44] : 터미널 서비스 암호화 수준 설정 : " + str(res.read())
                            self.status_code['W-44'] = "취약"
                        else:
                            result = "[W-44] : 터미널 서비스 암호화 수준 설정 : " + str(res.read())
                            self.status_code['W-44'] = "양호"
                    res.close()
                except:
                    result = "[W-44] : 터미널 서비스 암호화 수준 설정 : 최소 레벨 존재하지 않음[MinEncryptionLevel]"
                    res.close()
            else:
                result = "[W-44] : 터미널 서비스 암호화 수준 설정 : 사용 안 함"
                self.status_code['W-44'] = "양호"
            self.result_list['W-44'] = result
            res.close()
        except:
            result = "[W-44] : 터미널 서비스 암호화 수준 설정 : 알 수 없음"
            self.result_list['W-44'] = result
            self.status_code['W-44'] = "비 정 상"
            res.close()

    # IIS 웹 서비스 정보 숨김
    def w45(self):
        try:
            check = self.common_func("%systemroot%\system32\inetsrv\\appcmd list site /config | find /i \"physicalpath=\"")
            checks = check.read().count("\n")
            check.close()
            if checks >= 1:
                appends = ""
                for x in range(checks):
                    res = self.common_func("%systemroot%\system32\inetsrv\\appcmd list site /config | find /i \"physicalpath=\"")
                    result = re.search(re.compile("physicalPath=(\")(.*)(\" )"), res.read()).group(2)
                    appends = appends + " " + result
                    res.close()
                    if len(result) > 1:
                        result = self.common_func("type {}\web.config | find \"statuscode\" /i".format(result))
                        if len(result) > 1:
                            result = "[W-45] IIS 웹 서비스 정보 숨김 : " + result + "존재"
                            self.status_code['W-45'] = "양호"
            else:
                result = "[W-45] IIS 웹 서비스 정보 숨김 : 에러 페이지 처리가 존재하지 않음"
                self.status_code['W-45'] = "취약"
            self.result_list['W-45'] = result
        except:
            result = "[W-45] IIS 웹 서비스 정보 숨김 : 알 수 없음"
            self.status_code['W-45'] = "비 정 상"
            self.result_list['W-45'] = result

    # SNMP 서비스 구동 점검
    def w46(self):
        try:
            res = self.common_func("net start | find /i \"SNMP Service\"")
            if len(res.read()) > 1:
                result = "[W-46] SNMP 서비스 구동 점검 : " + str(res.read()) + " 사용 중"
                self.status_code['W-46'] = "취약"
            else:
                result = "[W-46] SNMP 서비스 구동 점검 : 사용 안 함"
                self.status_code['W-46'] = "양호"
            self.result_list['W-46'] = result
            res.close()
        except:
            result = "[W-46] SNMP 서비스 구동 점검 : 알 수 없음"
            self.result_list['W-46'] = result
            self.status_code['W-46'] = "비 정 상"
            res.close()

    # SNMP 서비스 Comunity String의 복잡성 설정
    def w47(self):
        try:
            result = Mygetreg.w47_registry()
            if result is not None:
                result2 = result[0] # public OR private
                result3 = result[1] # 0x 뒤 숫자
                permission_string = ["public","private","PUBLIC","PRIVATE"]
                permission_number = ["8,","16"]
                if result3 in permission_number:
                    self.status_code['W-47'] = "취약"
                    result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : " + str(result3) + " [읽기 이상의 권한]"
                elif result2 in permission_string:
                    self.status_code['W-47'] = "취약"
                    result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : " + str(result2) + " [읽기 이상의 권한]"
                elif result2 in permission_string and result3 in permission_number:
                    self.status_code['W-47'] = "취약"
                    result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : " + str(result2) + " / " + str(result3) +" - [읽기 이상의 권한]"
                else:
                    self.status_code['W-47'] = "양호"
                    result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : " + str(result)
                self.result_list['W-47'] = result
            else:
                self.status_code['W-47'] = "취약"
                result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : 설정 값 없음"
            self.result_list['W-47'] = result
        except:
            result = "[W-47] SNMP 서비스 커뮤니티 스트링의 복잡성 설정 : 알 수 없음"
            self.result_list['W-47'] = result
            self.status_code['W-47'] = "비 정 상"

    # SMP Acess control 설정
    def w48(self):
        try:
            result = Mygetreg.w48_registry()
            if result is not None:
                result = "[W-48] SNMP Access Control 설정 : " + str(result)
                self.status_code['W-48'] = "양호"
            else:
                result = "[W-48] SNMP Access Control 설정 : localhost 혹은 모든 호스트 허용"
                self.status_code['W-48'] = "취약"
            self.result_list['W-48'] = result
        except:
            result = "[W-48] SNMP Access Control 설정 : 알 수 없음"
            self.result_list['W-48'] = result
            self.status_code['W-48'] = "비 정 상"

    # DNS 서비스 구동 점검
    def w49(self):
        try:
            result = Mygetreg.w49_registry()
            if result is not None:
                if "0x1" in result:
                    self.status_code['W-49'] = "취약"
                    result = "[W-49] : DNS 서비스 구동 점검 : " + str(result) + "[보안 되지않음]"
                elif "0x0" in result:
                    self.status_code['W-49'] = "양호"
                    result = "[W-49] : DNS 서비스 구동 점검 : " + str(result) + "[동적 업데이트 없음]"
                else:
                    self.status_code['W-49'] = "양호"
                    result = "[W-49] : DNS 서비스 구동 점검 : " + str(result) + "[동적 업데이트 없음]"
            else:
                self.status_code['W-49'] = "양호"
                result = "[W-49] : DNS 서비스 구동 점검 : 파일이 존재하지 않음"
            self.result_list['W-49'] = result
        except:
            result = "[W-49] : DNS 서비스 구동 점검 : 알 수 없음"
            self.result_list['W-49'] = result
            self.status_code['W-49'] = "비 정 상"

    # HTTP/FTTP/SMTP 배너 차단
    def w50(self):
        try:
            ftp_result = self.common_func("type C:\Windows\System32\inetsrv\config\\applicationHost.config | find \"messages bannerMessage\"")
            result = "[W-50] HTTP/ FTP/ SMTP 배너 차단 : "
            comp = 0
            if len(ftp_result.read()) > 1:
                result = result +"\n"+ ftp_result.read()
                comp = comp+1
            ftp_result.close()
            smtp_result = self.common_func("type C:\Windows\System32\inetsrv\MetaBase.xml | find \"ConnectResponse\"")
            if len(smtp_result.read()) > 1:
                result = result + "\n" + smtp_result.read()
                comp = comp + 1
            smtp_result.close()
            http_result = telnet_banner.http_banner()
            if "" not in http_result:
                result = result + "\n" + http_result
                comp = comp + 1
            if comp is not 0:
                self.status_code['W-50'] = "취약"
            else:
                self.status_code['W-50'] = "양호"
            self.result_list['W-50'] = result
        except:
            result = "[W-50] HTTP/ FTP/ SMTP 배너 차단 : 알 수 없음"
            self.result_list['W-50'] = result
            self.status_code['W-50'] = "비 정 상"

    # Telnet 보안 설정
    def w51(self):
        try:
            result = platform.release()
            if result == "10":
                self.status_code['W-51'] = "양호"
                result = "[W-51] : Telnet 보안 설정 : Windows10은 Telnet Service를 제공하지 않습니다."
            else:
                result = Mygetreg.w51_registry()
                if result is not None:
                    if int(result) == 2:
                        self.status_code['W-51'] = "양호"
                        result = "[W-51] : Telnet 보안 설정 : NTLM 인증 사용"
                    else:
                        self.status_code['W-51'] = "취약"
                        result = "[W-51] : Telnet 보안 설정 : NTLM 인증을 사용하지 않음"
                else:
                    result = "[W-51] : Telnet 보안 설정 : Telnet Server가 존재하지 않습니다."
                    self.status_code['W-51'] = "양호"
            self.result_list['W-51'] = result
        except:
            result = "[W-51] : Telnet 보안 설정 : 알 수 없음"
            self.result_list['W-51'] = result
            self.status_code['W-51'] = "비 정 상"

    # 불 필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거
    def w52(self):
        try:
            result = Mygetreg.odbc()
            if result == "ODBC Data Sources":
                self.status_code['W-52'] = "양호"
                result = "[W-52] 불 필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거 : " + str(result) +"[사용중]"
            else:
                self.status_code['W-52'] = "취약"
                result = "[W-52] 불 필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거 : 사용 안 함"
            self.result_list['W-52'] = result
        except:
            result = "[W-52] 불 필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거 : 알 수 없음"
            self.status_code['W-52'] = "비 정 상"
            self.result_list['W-52'] = result

    # 원격 터미널 접속 타임아웃 설정
    def w53(self):
        try:
            result = Mygetreg.w53_registry() # tuple(1,[0,0,0])
            if result[0] is True:
                result = result[1].get('MaxIdleTime')
                if result >= 3600000:
                    result = result / 1000 / 3600 # 시간단위 계산
                    result = "[W-53] 원격 터미널 접속 타임아웃 설정 : " + str(int(result)) + "시간"
                else:
                    result = result / 1000 / 60 # 분 단위 계산
                    result = "[W-53] 원격 터미널 접속 타임아웃 설정 : " + str(int(result)) + "분"
                self.status_code['W-53'] = "양호"
            else:
                result = "[W-53] 원격 터미널 접속 타임아웃 설정 : 구성되지 않음"
                self.status_code['W-53'] = "취약"
            self.result_list['W-53'] = result
        except:
            result = "[W-53] 원격 터미널 접속 타임아웃 설정 : 알 수 없음"
            self.result_list['W-53'] = result
            self.status_code['W-53'] = "비 정 상"

    # 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검
    def w54(self):
        try:
            result = self.file_search()
            if len(result) > 1:
                self.status_code['W-54'] = "취약"
                result = "[W-54] 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검 : " + "\n" + str(result)
            else:
                result = "[W-54] 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검 : 예약 작업이 존재하지 않음"
                self.status_code['W-54'] = "양호"
            self.result_list['W-54'] = result
        except:
            result = "[W-54] 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검 : 알 수 없음"
            self.result_list['W-54'] = result
            self.status_code['W-54'] = "비 정 상"

     ####################################### 패치 관리 ####################################

    # 최신 HOT FIX 적용
    def w55(self):
        try:
            result = hotfix.hotfix_comp()
            if "누락" in result:
                self.status_code['W-55'] = "취약"
                result = "[W-55] 최신 HOT FIX 적용 : " + str(result)
            elif "완료" in result:
                self.status_code['W-55'] = "양호"
                result = "[W-55] 최신 HOT FIX 적용 : " + str(result)
            else:
                self.status_code['W-55'] = "비 정 상"
                result = "[W-55] 최신 HOT FIX 적용 : 알 수 없음"
            self.result_list['W-55'] = result
        except:
            result = "[W-55] 최신 HOT FIX 적용 : 알 수 없음"
            self.result_list['W-55'] = result
            self.status_code['W-55'] = "비 정 상"

    # 백신 프로그램 업데이트
    def w56(self):
        try:
            result = findvacine.windows_defender()
            result = "[W-56] 백신 프로그램 업데이트 : [Windows Defender Version] = " + str(result) +"[수동비교 권장]"
            self.status_code['W-56'] = "수동"
            self.result_list['W-56'] = result
        except:
            result = "[W-56] 백신 프로그램 업데이트 : 알 수 없음"
            self.status_code['W-56'] = "비 정 상"
            self.result_list['W-56'] = result

    # 정책에 따른 시스템 로깅 설정
    def w57(self):
        try:
            with open("Local_Security_Policy.txt", "r",encoding="utf-16") as f:
                result = f.read()
                result1 = re.search(re.compile("AuditLogonEvents = ([0-9])"),result).group(1) # 로그온 이벤트
                result2 = re.search(re.compile("AuditPrivilegeUse = ([0-9])"),result).group(1) #권한 사용
                result3 = re.search(re.compile("AuditAccountManage = ([0-9])"),result).group(1) # 계정 관리
                result4 = re.search(re.compile("AuditPolicyChange = ([0-9])"),result).group(1) # 정책 변경
                result5 = re.search(re.compile("AuditDSAccess = ([0-9])"),result).group(1) # 디렉터리 액세스
                result6 = re.search(re.compile("AuditAccountLogon = ([0-9])"), result).group(1)  # 계정 로그온 이벤트
            if "0" in result1 or "0" in result2 or "0" in result3 or "0" in result4 or "0" in result5 or "0" in result6:
                self.status_code['W-57'] = "취약"
                result = "[W-57] 정책에 따른 시스템 로깅 설정 : \n로그온 이벤트 : " + result1 +"\n계정 로그온 이벤트 : " + result6 + "\n정책 변경 : " + result4 + "\n계정 관리 : "+result3 +"\n디렉터리 액세스 : "+result5 +"\n권한 사용 : "+ result2

            elif "1" in result1 or "1" in result2 or "1" in result3 or "1" in result4 or "1" in result5 or "1" in result6:
                result = "[W-57] 정책에 따른 시스템 로깅 설정 : \n로그온 이벤트 : " + result1 + "\n계정 로그온 이벤트 : " + result6 + "\n정책 변경 : " + result4 + "\n계정 관리 : " + result3 + "\n디렉터리 액세스 : " + result5 + "\n권한 사용 : " + result2
                if "3" in result1 and "3" in result6 and "3" in result4 and "2" in result2 and "2" in result3 and "2" in result5:
                    self.status_code['W-57'] = "양호"
                else:
                    self.status_code['W-57'] = "취약"

            elif "2" in result1 or "2" in result2 or "2" in result3 or "2" in result4 or "2" in result5 or "2" in result6:
                result = "[W-57] 정책에 따른 시스템 로깅 설정 : \n로그온 이벤트 : " + result1 + "\n계정 로그온 이벤트 : " + result6 + "\n정책 변경 : " + result4 + "\n계정 관리 : " + result3 + "\n디렉터리 액세스 : " + result5 + "\n권한 사용 : " + result2
                if "3" in result1 and "3" in result6 and "3" in result4 and "2" in result2 and "2" in result3 and "2" in result5:
                    self.status_code['W-57'] = "양호"
                else:
                    self.status_code['W-57'] = "취약"

            elif "3" in result1 or "3" in result2 or "3" in result3 or "3" in result4 or "3" in result5 or "3" in result6:
                result = "[W-57] 정책에 따른 시스템 로깅 설정 : \n로그온 이벤트 : " + result1 + "\n계정 로그온 이벤트 : " + result6 + "\n정책 변경 : " + result4 + "\n계정 관리 : " + result3 + "\n디렉터리 액세스 : " + result5 + "\n권한 사용 : " + result2
                if "3" in result1 and "3" in result6 and "3" in result4 and "2" in result2 and "2" in result3 and "2" in result5:
                    self.status_code['W-57'] = "양호"
                else:
                    self.status_code['W-57'] = "취약"
            else:
                self.status_code['W-57'] = "비 정 상"
            self.result_list['W-57'] = result
        except:
            result = "[W-57] 정책에 따른 시스템 로깅 설정 : 알 수 없음"
            self.result_list['W-57'] = result
            self.status_code['W-57'] = "비 정 상"

    ####################################### 로그 관리 ####################################

    # 로그의 정기적 검토 및 보고 - 수동점검
    def w58(self):
        try:
            result =  "[W-58] 로그의 정기적 검토 및 보고 : 수동"
            self.result_list['W-58'] = result
            self.status_code['W-58'] = "수동"
        except:
            result = "[W-58] 로그의 정기적 검토 및 보고 : 알 수 없음"
            self.result_list['W-58'] = result
            self.status_code['W-58'] = "비 정 상"

    # 원격으로 액세스 할 수 있는 레지스트리 경로
    def w59(self):
        try:
            res = self.common_func("net start | find /i \"Remote Registry\"")
            if len(res.read()) > 1:
                self.status_code['W-59'] = "취약"
                result = "[W-59] 원격으로 액세스 할 수 있는 레지스트리 경로 : \n" + str(res.read())
            else:
                self.status_code['W-59'] = "양호"
                result = "[W-59] 원격으로 액세스 할 수 있는 레지스트리 경로 : 없음"
            self.result_list['W-59'] = result
        except:
            result = "[W-59] 원격으로 액세스 할 수 있는 레지스트리 경로 : 알 수 없음"
            self.result_list['W-59'] = result
            self.status_code['W-59'] = "비 정 상"

    # 이벤트 로그 관리 설정
    def w60(self):
        try:
            result = Mygetreg.w60_registry()
            result1 = result[0][0]
            result2 = result[1][0]
            result3 = result[2][0]
            try:
                result4 = result[0][1]
            except IndexError:
                result4 = 0 # AutoBackupLogoFiles가 존재하지않아서 에러발생 하는 경우 설정 값이 존재하지 않으므로 취약으로 진단하기 위해 '0' 값으로 설정
            try:
                result5 = result[1][1]
            except IndexError:
                result5 = 0 # AutoBackupLogoFiles가 존재하지않아서 에러발생 하는 경우 설정 값이 존재하지 않으므로 취약으로 진단하기 위해 '0' 값으로 설정
            try:
                result6 = result[2][1]
            except IndexError:
                result6 = 0 # AutoBackupLogoFiles가 존재하지않아서 에러발생 하는 경우 설정 값이 존재하지 않으므로 취약으로 진단하기 위해 '0' 값으로 설정
            result = "[W-60] 이벤트 로그 관리 설정 : "
            if result1 > 10240 or result2 > 10240 or result3 > 10240 or result4 < 1 or result5 < 1 or result6 < 1:
                self.status_code['W-60'] = "취약"
                result = result + "\nApplication Log 최대 크기 : " + str(result1) + "\nSecurity Log 최대 크기 : " + str(result2) + "\nSystem Log 최대 크기 : " +str(result3)
                result = result + "\nApplication " + str(result4) + " 일 이후 자동 덮어쓰기 설정\nSecurity " + str(result5) + " 일 이후 자동 덮어쓰기 설정\nSystem " + str(result6) + " 일 이후 자동 덮어쓰기 설정"
            else:
                self.status_code['W-60'] = "양호"
                result = result + "\nApplication Log 최대 크기 : " + str(result1) + "\nSecurity Log 최대 크기 : " + str(result2) + "\nSystem Log 최대 크기 : " + str(result3)
                result = result + "\nApplication " + str(result4) + " 일 이후 자동 덮어쓰기 설정\nSecurity " + str(result5) + " 일 이후 자동 덮어쓰기 설정\nSystem " + str(result6) + " 일 이후 자동 덮어쓰기 설정"
            self.result_list['W-60'] = result
        except:
            result = "[W-60] 이벤트 로그 관리 설정 : 알 수 없음"
            self.result_list['W-60'] = result
            self.status_code['W-60'] = "비 정 상"

    # 원격에서 이벤트 로그 파일 접근 차단
    def w61(self):
        try:
            res1 = self.common_func("cacls %systemroot%\system32\logfiles | find /i \"everyone\"")
            res2 = self.common_func("cacls %systemroot%\system32\config | find /i \"everyone\"")
            if len(res1.read()) > 1 or len(res2.read()) > 1:
                result = "[W-61] 원격에서 이벤트 로그 파일 접근 차단 : \n"+ res1.read() +"\n"+ res2.read()
                self.status_code['W-61'] = "취약"
            else:
                result = "[W-61] 원격에서 이벤트 로그 파일 접근 차단 : Log Directory의 Everyone 권한 없음"
                self.status_code['W-61'] = "양호"
            self.result_list['W-61'] = result
            res1.close()
            res2.close()
        except:
            result = "[W-61] 원격에서 이벤트 로그 파일 접근 차단 : 알 수 없음"
            self.result_list['W-61'] = result
            self.status_code['W-61'] = "취약"

    # 백신 프로그램 설치
    def w62(self):
        res = self.common_func("net start | find /i \"Windows Defender Service\"")
        try:
            if len(res.read()) > 1:
                result = "[W-62] 백신 프로그램 설치 : " + str(res.read()) + "[동작중]"
                self.status_code['W-62'] = "양호"
            else:
                result = "[W-62] 백신 프로그램 설치 : Windows Defender 동작 안 함"
                self.status_code['W-62'] = "취약"
            self.result_list['W-62'] = result
            res.close()
        except:
            result = "[W-62] 백신 프로그램 설치 : 알 수 없음"
            self.result_list['W-62'] = result
            self.status_code['W-62'] = "비 정 상"
            res.close()

    #  SAM 파일 접근 통제 설정
    def w63(self):
        res = self.common_func("cacls %systemroot%\system32\config\SAM | findstr /v \"Administrator SYSTEM\" | findstr \"\\ : \"")
        try:
            if len(res.read()) > 1:
                result = "[W-63] SAM 파일 접근 통제 설정: " + str(res.read())
                self.status_code['W-63'] = "취약"
            else:
                result = "[W-63] SAM 파일 접근 통제 설정: 없음"
                self.status_code['W-63'] = "양호"
            self.result_list['W-63'] = result
            res.close()
        except:
            result = "[W-63] SAM 파일 접근 통제 설정: 알 수 없음"
            self.result_list['W-63'] = result
            self.status_code['W-63'] = "비 정 상"
            res.close()

    # 화면 보호기 설정
    def w64(self):
        try:
            result = Mygetreg.w64_registry()
            if result[0] == "1" and len(result) > 1:
                if int(result[1]) >= 600 and result[2] == "1":
                    self.status_code['W-64'] = "양호"
                    result = "[W-64] 화면 보호기 설정 : \n화면 보호기 활성 : True\n화면 보호기 시간 제한 : " + str(result[1]) + "\n화면 보호기 해제 시 로그온 유무 : True"
                else:
                    self.status_code['W-64'] = "취약"
                    if result[2] == "0":
                        result = "[W-64] 화면 보호기 설정 : \n화면 보호기 활성 : True\n화면 보호기 시간 제한 : " + str(result[1]) + "\n화면 보호기 해제 시 로그온 유무 : False"
                    else:
                        result = "[W-64] 화면 보호기 설정 : \n화면 보호기 활성 : True\n화면 보호기 시간 제한 : " + str(result[1]) + "\n화면 보호기 해제 시 로그온 유무 : True"
            else:
                self.status_code['W-64'] = "취약"
                result = "[W-64] 화면 보호기 설정 : 비 활성화"
            self.result_list['W-64'] = result
        except:
            result = "[W-64] 화면 보호기 설정 : 알 수 없음"
            self.result_list['W-64'] = result
            self.status_code['W-64'] = "비 정 상"

    # 로그온 하지 않고 시스템 종료 허용
    def w65(self):
        try:
            result = Mygetreg.w65_registry()
            if result == 1:
                self.status_code['W-65'] = "양호"
                result = "[W-65] 로그온 하지 않고 시스템 종료 허용 : 허용"
            elif result == 0:
                self.status_code['W-65'] = "취약"
                result = "[W-65] 로그온 하지 않고 시스템 종료 허용 : 허용 안 함"
            else:
                self.status_code['W-65'] = "비 정 상"
                result = "[W-65] 로그온 하지 않고 시스템 종료 허용 : 설정 오류"
            self.result_list['W-65'] = result
        except:
            result = "[W-65] 로그온 하지 않고 시스템 종료 허용 : 알 수 없음"
            self.result_list['W-65'] = result
            self.status_code['W-65'] = "비 정 상"

    # 원격 시스템 강제로 시스템 종료
    def w66(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("SeRemoteShutdownPrivilege =\s+(.*)"),result).group(1).replace("*S-1-5-32-544","")
            if len(result) > 1:
                self.status_code['W-66'] = "취약"
                result = "[W-66] 원격 시스템 강제로 시스템 종료 : " + str(result)
            else:
                result = "[W-66] 원격 시스템 강제로 시스템 종료 : Administrators"
                self.status_code['W-66'] = "양호"
            self.result_list['W-66'] = result
        except:
            result = "[W-66] 원격 시스템 강제로 시스템 종료 : 알 수 없음"
            self.result_list['W-66'] = result
            self.status_code['W-66'] = "비 정 상"

    # 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료
    def w67(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("CrashOnAuditFail=([0-9].[0-9])"), result).group(1)
            if result == "4,0":
                result = "[W-67] 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료 : 즉시종료 안 함"
                self.status_code['W-67'] = "양호"
            elif result == "4,1":
                result = "[W-67] 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료 : 즉시종료 함"
                self.status_code['W-67'] = "취약"
            else:
                result = "[W-67] 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료 : 알 수 없음"
                self.status_code['W-67'] = "비 정 상"
            self.result_list['W-67'] = result
        except:
            result = "[W-67] 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료 : 알 수 없음"
            self.result_list['W-67'] = result
            self.status_code['W-67'] = "비 정 상"

    # SAM 계정과 공유의 익명 열거 허용 안 함
    def w68(self):
        try:
            result = Mygetreg.w68_registry()
            if result == 1:
                self.status_code['W-68'] = "양호"
                result = "[W-68] SAM 계정과 공유의 익명 열거 허용 안함 : 허용 안 함"
            elif result == 0:
                self.status_code['W-68'] = "취약"
                result = "[W-68] SAM 계정과 공유의 익명 열거 허용 안함 : 허용"
            else:
                result = "[W-68] SAM 계정과 공유의 익명 열거 허용 안함 : 설정 값 오류"
                self.status_code['W-68'] = "비 정 상"
            self.result_list['W-68'] = result
        except:
            result = "[W-68] SAM 계정과 공유의 익명 열거 허용 안함 : 알 수 없음"
            self.result_list['W-68'] = result
            self.status_code['W-68'] = "비 정 상"

    # Autologon 기능 제어
    def w69(self):
        try:
            result = Mygetreg.w69_registry()
            if result == "0":
                self.status_code['W-69'] = "양호"
                result = "[W-69] AutoLogon 기능 제어 : 사용 안 함"
            elif result == "1":
                self.status_code['W-69'] = "취약"
                result = "[W-69] AutoLogon 기능 제어 : 사용 안 함"
            elif result == None:
                self.status_code['W-69'] = "취약"
                result = "[W-69] AutoLogon 기능 제어 : 사용 안 함"
            else:
                self.status_code['W-69'] = "비 정 상"
                result = "[W-69] AutoLogon 기능 제어 : 설정 값 오류"
            self.result_list['W-69'] = result
        except:
            result = "[W-69] AutoLogon 기능 제어 : 알 수 없음"
            self.result_list['W-69'] = result
            self.status_code['W-69'] = "비 정 상"

    # 이동식 미디어 포맷 및 꺼내기 허용
    def w70(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                try:
                    result = re.search(re.compile("AllocateDASD=1,(.*)"),result).group(1)
                except AttributeError:
                    result = "설정값 없음"
            if "0" in str(result):
                self.status_code['W-70'] = "양호"
                result = "[W-70] 이동식 미디어 포맷 및 꺼내기 허용 : Administrators"
            elif "1" in str(result):
                self.status_code['W-70'] = "취약"
                result = "[W-70] 이동식 미디어 포맷 및 꺼내기 허용 : Administrators 및 Power Users"
            elif "2" in str(result):
                self.status_code['W-70'] = "취약"
                result = "[W-70] 이동식 미디어 포맷 및 꺼내기 허용 : Administrators 및 Interactive Users"
            else:
                self.status_code['W-70'] = "취약"
                result = "[W-70] 이동식 미디어 포맷 및 꺼내기 허용 : 설정 값 없음"
            self.result_list['W-70'] = result
        except:
            result = "[W-70] 이동식 미디어 포맷 및 꺼내기 허용 : 알 수 없음"
            self.result_list['W-70'] = result
            self.status_code['W-70'] = "비 정 상"

    # 디스크볼륨 암호화 설정 -- 측정불가
    def w71(self):
        try:
            result = "[W-71] 디스크 볼륨 암호화 설정 : 검사대상의 양이 방대하여 검사 불가"
            self.result_list['W-71'] = result
            self.status_code['W-71'] = "측정불가"
        except:
            result = "[W-71] 디스크 볼륨 암호화 설정 : 검사대상의 양이 방대하여 검사불가"
            self.result_list['W-71'] = result
            self.status_code['W-71'] = "측정불가"

    # Dos공격 방어 레지스트리 설정
    def w72(self):
        try:
            result = Mygetreg.w72_registry()
            if len(result) > 1:
                try:
                    if int(result[0],16) <= 1 and int(result[1],16) == 0 and int(result[2],16) <=300000 and int(result[3],16) == 1:
                        self.status_code['W-72'] = "양호"
                        result = "[W-72] Dos공격 방어 레지스트리 설정 : \nSynAttackProtect : " + str(result[0]) + "\nEnableDeadGWDetect : "+ str(result[1]) + "\nKeepAliveTime : " + str(result[2]) +"\nNoNameReleaseOnDemand : "+ str(result[3])
                    else:
                        self.status_code['W-72'] = "취약"
                        result = "[W-72] Dos공격 방어 레지스트리 설정 : \nSynAttackProtect : " + str(result[0]) + "\nEnableDeadGWDetect : " + str(result[1]) + "\nKeepAliveTime : " + str(result[2]) + "\nNoNameReleaseOnDemand : " + str(result[3])
                except:
                    result = "[W-72] Dos공격 방어 레지스트리 설정 : 4가지 설정 키중 존재하지 않는 키가 존재합니다."
                    self.status_code['W-72'] = "취약"
            else:
                result = "[W-72] Dos공격 방어 레지스트리 설정 : " + platform.system() +" " + platform.version() + " 은 안전한 Version 입니다."
                self.status_code['W-72'] = "양호"
            self.result_list['W-72'] = result
        except:
            result = "[W-72] Dos공격 방어 레지스트리 설정 : 알 수 없음"
            self.result_list['W-72'] = result
            self.status_code['W-72'] = "비 정 상"

    # 사용자가 프린터 드라이버를 설치할 수 없게 함
    def w73(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("AddPrinterDrivers=4,([0-9])"), result).group(1)
            if type(result) is not int:
                self.status_code['W-73'] = "취약"
                result = "[W-73] 사용자가 프린터 드라이버를 설치할 수 없게 함 : 설정 값 없음"
            elif "0" in result:
                self.status_code['W-73'] = "취약"
                result = "[W-73] 사용자가 프린터 드라이버를 설치할 수 없게 함 : 사용 안 함"
            elif "1" in result:
                self.status_code['W-73'] = "양호"
                result = "[W-73] 사용자가 프린터 드라이버를 설치할 수 없게 함 : 사용 함"
            else:
                self.status_code['W-73'] = "취약"
                result = "[W-73] 사용자가 프린터 드라이버를 설치할 수 없게 함 : 설정 값 없음"
            self.result_list['W-73'] = result
        except:
            result = "[W-73] 사용자가 프린터 드라이버를 설치할 수 없게 함 : 알 수 없음"
            self.result_list['W-73'] = result
            self.status_code['W-73'] = "비 정 상"

    # 세션 연결을 중단하기 전에 필요한 유휴 시간
    def w74(self):
        result = Mygetreg.w74_registry()
        try:
            if result.get('EnableForcedLogoff') == 1:
                if len(result) == 1:
                    self.status_code['W-74'] = "취약"
                    result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : 설정 값 없음"
                elif result.get('AutoDisconnect') >= 15:
                    self.status_code['W-74'] = "양호"
                    result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : " + str(result.get('AutoDisconnect')) + "분"
                else:
                    self.status_code['W-74'] = "취약"
                    result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : " + str(result.get('AutoDisconnect')) + "분"
            elif result.get('EnableForcedLogoff') == 0:
                self.status_code['W-74'] = "취약"
                result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : 사용 안 함"
            else:
                self.status_code['W-74'] = "취약"
                result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : 설정 값 없음"
            self.result_list['W-74'] = result
        except:
            result = "[W-74] 세션 연결을 중단하기 전에 필요한 유휴 시간 : 알 수 없음"
            self.status_code['W-74'] = "비 정 상"
            self.result_list['W-74'] = result

    # 경고 메시지 설정
    def w75(self):
        try:
            result = Mygetreg.w75_registry()
            result1 = result.get('LegalNoticeCaption')
            result2 = result.get('LegalNoticeText')
            if len(result1) or len(result2) > 1:
                self.status_code['W-75'] = "양호"
                result = "[W-75] 경고 메시지 설정 : \n" + str(result) + "\n" + str(result)
            else:
                self.status_code['W-75'] = "취약"
                result = "[W-75] 경고 메시지 설정 : 없음"
            self.result_list['W-75'] = result
        except:
            result = "[W-75] 경고 메시지 설정 : 알 수 없음"
            self.result_list['W-75'] = result
            self.status_code['W-75'] = "비 정 상"

    # 사용자 별 홈 디렉터리 권한 설정
    def w76(self):
        try:
            value = self.common_func("dir \"{}\\Users\\*\" | find \"<DIR>\"  | findstr /V \"Public All Default .\"".format(os.getenv("SystemDrive")))
            values = value.read().split("\n")
            value.close()
            for var in values:
                if var == "":
                    break
                username = re.search(re.compile("<DIR>\s+(.*)"), var).group(1)
                res = self.common_func("icacls \"{}\\Documents and Settings\\{}\" | find /i \"everyone\"".format(os.getenv("SystemDrive"),username))
                if len(res.read()) > 1:
                    self.status_code['W-76'] = "취약"
                    result = "[W-76] 사용자 별 홈 디렉터리 권한 설정 : " + str(res.read()) + " [Everyone 권한 존재]"
                    res.close()
                    break
                else:
                    self.status_code['W-76'] = "양호"
                    result = "[W-76] 사용자 별 홈 디렉터리 권한 설정 : Everyone 권한 없음"
                    res.close()
                    break
            self.result_list['W-76'] = result

        except:
            result = "[W-76] 사용자 별 홈 디렉터리 권한 설정 : 알 수 없음"
            self.result_list['W-76'] = result
            self.status_code['W-76'] = "비 정 상"

    # LAN Manager 인증 수준
    def w77(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result = re.search(re.compile("LmCompatibilityLevel=4,([0-9])"), result).group(1)
            if "0" in result:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : LM 및 NTLM 응답 보내기"
            elif "1" in result:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : LM 및 NTLM 보내기 - 협상되면 NTLMv2 세션 보안 사용(&)"
            elif "2" in result:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : NTLM 응답만 보내기"
            elif "3" in result:
                self.status_code['W-77'] = "양호"
                result = "[W-77] LAN Manager 인증 수준 : NTLMv2 응답만 보내기"
            elif "4" in result:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : NTLM 응답만 보내기 및 LM 거부"
            elif "5" in result:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : NTLMv2 응답만 보냅니다. LM 및 NTLM은 거부합니다."
            else:
                self.status_code['W-77'] = "취약"
                result = "[W-77] LAN Manager 인증 수준 : 설정 값 없음"
            self.result_list['W-77'] = result
        except AttributeError:
            result = "[W-77] LAN Manager 인증 수준 : 설정 값 없음"
            self.status_code['W-77'] = "취약"
            self.result_list['W-77'] = result
        except:
            result = "[W-77] LAN Manager 인증 수준 : 알 수 없음"
            self.status_code['W-77'] = "비 정 상"
            self.result_list['W-77'] = result

    # 보안 채널 데이터 디지털 암호화 또는 서명
    def w78(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result1 = re.search(re.compile("RequireSignOrSeal=4,([0-9])"), result).group(1)
                result2 = re.search(re.compile("SealSecureChannel=4,([0-9])"), result).group(1)
                result3 = re.search(re.compile("SignSecureChannel=4,([0-9])"), result).group(1)
                result = "[W-78] 보안 채널 데이터 디지털 암호화 또는 서명 : \n"
                if result1 == 0:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 서명(가능한 경우) [사용 안 함]\n"
                else:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 서명(가능한 경우) [사용 함]\n"

                if result2 == 0:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 암호화 또는 서명(항상) [사용 안 함]\n"
                else:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 암호화 또는 서명(항상) [사용]\n"

                if result3 == 0:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 암호화(가능한 경우) [사용 안 함]"
                else:
                    result = result + "도메인 구성원 : 보안 채널 데이터를 디지털 암호화(가능한 경우) [사용 함]"

                if int(result1) >= 1 and int(result2) >= 1 and int(result3) >= 3:
                    self.status_code['W-78'] = "양호"
                else:
                    self.status_code['W-78'] = "취약"
                self.result_list['W-78'] = result
        except:
            result = "[W-78] 보안 채널 데이터 디지털 암호화 또는 서명 : 알 수 없음"
            self.result_list['W-78'] = result
            self.status_code['W-78'] = "비 정 상"

    # 파일 및 디렉터리 보호
    def w79(self):
        try:
            # res = self.common_func("fsutil fsinfo volumeinfo {} | find /i \"File System Name\"".format(os.getenv("SystemDrive")))
            res = winstats.get_vol_info(os.getenv("SystemDrive").replace(":","")).fstype
            if "NTFS" in res:
                self.status_code['W-79'] = "양호"
                result = "[W-79] 파일 및 디렉터리 보호 : \nFile System : " + str(res)
            elif "FAT" in res:
                self.status_code['W-79'] = "취약"
                result = "[W-79] 파일 및 디렉터리 보호 : \nFile System : " + str(res)
            else:
                self.status_code['W-79'] = "취약"
                result = "[W-79] 파일 및 디렉터리 보호 : \nFile System : " + str(res)
            self.result_list['W-79'] = result
        except AttributeError:
            result = "[W-79] 파일 및 디렉터리 보호 : 설정 값 오류"
            self.result_list['W-79'] = result
            self.status_code['W-79'] = "취약"
        except:
            result = "[W-79] 파일 및 디렉터리 보호 : 알 수 없음"
            self.result_list['W-79'] = result
            self.status_code['W-79'] = "비 정 상"

    # 컴퓨터 계정 암호 최대 사용 기간
    def w80(self):
        try:
            with open("Local_Security_Policy.txt","r",encoding="utf-16") as f:
                result = f.read()
                result1 = re.search(re.compile("DisablePasswordChange=4,([0-9])"), result).group(1)
                result2 = re.search(re.compile("aximumPasswordAge=4,([0-9])"), result).group(1)
            result = "[W-80] 컴퓨터 계정 암호 최대 사용 기간 : {}일\n".format(result2)
            if int(result1) == 0 and int(result2) >= 90:
                self.status_code['W-80'] = "양호"
                result = str(result) + "컴퓨터 계정 암호 변경 사용 안 함 : 사용 안 함"
            elif int(result1) == 1 or int(result2) < 90:
                self.status_code['W-80'] = "취약"
                result = str(result) + "컴퓨터 계정 암호 변경 사용 안 함 : 사용 함"
            else:
                self.status_code['W-80'] = "양호"
                result = str(result) + "컴퓨터 계정 암호 변경 사용 안 함 : 설정 값 없음"
            self.result_list['W-80'] = result
        except AttributeError:
            result = "[W-80] 컴퓨터 계정 암호 최대 사용 기간 : 설정 값 없음"
            self.result_list['W-80'] = result
            self.status_code['W-80'] = "취약"
        except:
            result = "[W-80] 컴퓨터 계정 암호 최대 사용 기간 : 알 수 없음"
            self.result_list['W-80'] = result
            self.status_code['W-80'] = "비 정 상"

    # 시작프로그램 목록 분석
    def w81(self):
        try:
            username = os.getlogin()
            result1 = self.startup_D(username)
            result2 = self.startUp_files()
            result1 = [x for x in result1 if x not in result2]
            result2 = [x for x in result2 if x not in result1]
            result = result1 + result2
            if len(result) > 1:
                result = "[W-81] 시작 프로그램 목록 분석 : " + str(result).replace("[","").replace("]","").replace("\'","")
                self.status_code['W-81'] = "취약"
            else:
                result = "[W-81] 시작 프로그램 목록 분석 : 없음"
                self.status_code['W-81'] = "양호"
            self.result_list['W-81'] = result
        except:
            result = "[W-81] 시작 프로그램 목록 분석 : 알 수 없음"
            self.result_list['W-81'] = result
            self.status_code['W-81'] = "비 정 상"

    # Windows 인증 모드 사용
    def w82(self):
        try:
            res = Mygetreg.w82_registry()
            if res is not None:
                if str(res) == "2":
                    result = "[W-82] Windows 인증 모드 사용 : 혼합 인증 모드 사용"
                    self.result_list['W-82'] = result
                    self.status_code['W-82'] = '취약'
                elif str(res) == "1":
                    result = "[W-82] Windows 인증 모드 사용 : Windows 인증 모드 사용"
                    self.result_list['W-82'] = result
                    self.status_code['W-82'] = '양호'
                else:
                    result = "[W-82] Windows 인증 모드 사용 : 알 수 없음"
                    self.result_list['W-82'] = result
                    self.status_code['W-82'] = '비 정 상'
            else:
                result = "[W-82] Windows 인증 모드 사용 : MS SQL Server를 사용하고 있지 않습니다. \n(설정 값 없음)"
                self.result_list['W-82'] = result
                self.status_code['W-82'] = '양호'
        except:
            result = "[W-82] Windows 인증 모드 사용 : 알 수 없음"
            self.result_list['W-82'] = result
            self.status_code['W-82'] = '비 정 상'

class Check_Win(Check_Vulnerability_Exposures):

    def __init__(self):
        self.check = Mygetreg.iis_Check()
        self.ftp_comp = self.service_status("FTP")
        self.now = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')

    def start(self):
        if self.check is not None:
            self.run_Iis()
        else:
            self.is_not_Running_IIS()

    def is_not_Running_IIS(self):
        cw = Check_Vulnerability_Exposures()
        cw.w01()  # Administrator 계정 이름 바꾸기 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w02()  # Guest 계정 상태 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w03()  # 불필요한 계정 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w04()  # 계정 잠금 임계값 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w05()  # 해독 가능한 암호화를 사용하여 암호 저장 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w06()  # 관리자 그룹에 최소한의 사용자 포함 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w07()  # Everyone 사용 권한을 익명 사용자에게 적용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w08()  # 계정 잠금 기간 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w09()  # 패스워드 복잡성 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w10()  # 패스워드 최소 암호 길이 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w11()  # 패스워드 최대 사용 기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w12()  # 패스워드최소사용기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w13()  # 마지막 사용자 이름 표시 안함 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w14()  # 로컬 로그온 허용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w15()  # 익명 SID/이름 변환 허용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w16()  # 최근 암호 기억 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w17()  # 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w18()  # 원격터미널 접속 가능한 사용자 그룹 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w19()  # 공유 권한 및 사용자 그룹 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w20()  # 하드디스크 기본 공유 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w21()  # 불필요한 서비스 제거 / AttributeError 적용 완료 2017-11-17 (금)
        # iis 22 ~ 35 / 42 / 45 ~ 46 / 50
        for number in range(22,36):
            self.iis_not_Remainder(number,cw)
        # IIS 22 ~ 35
        cw.w36()  # NetBIOS 바인딩 서비스 구동 점검
        # FTP 37 ~ 40
        if self.ftp_comp is True:
            cw.w37()  # FTP 서비스 구동 점검
            cw.w38()  # FTP 디렉토리 접근권한 설정
            cw.w39()  # Anonymouse FTP 금지
            cw.w40()  # FTP 접근 제어 설정
        else:
            for ftp_numb in range(37,41):
                self.ftp_not_Remainder(ftp_numb,cw)
        cw.w41()  # DNS Zone Transfer 설정
        self.iis_not_Remainder(42, cw)  # IIS 42
        cw.w43()  # 최신 서비스팩 적용
        cw.w44()  # 터미널 서비스 암호화 수준 설정
        self.iis_not_Remainder(45, cw) # IIS 45
        self.iis_not_Remainder(46, cw) # IIS 46
        cw.w47()  # SNMP 서비스 커뮤니티스트링의 복잡성 설정
        cw.w48()  # SNMP Access control 설정
        cw.w49()  # DNS 서비스 구동 점검
        self.iis_not_Remainder(50, cw) # IIS 50
        cw.w51()  # Telnet 보안 설정
        cw.w52()  # 불필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거
        cw.w53()  # 원격터미널 접속 타임아웃 설정
        cw.w54()  # 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검
        cw.w55()  # 최신 HOT FIX 적용
        cw.w56()  # 백신 프로그램 업데이트
        cw.w57()  # 정책에 따른 시스템 로깅 설정
        cw.w58()  # 로그의 정기적 검토 및 보고
        cw.w59()  # 원격으로 액세스할 수 있는 레지스트리 경로
        cw.w60()  # 이벤트 로그 관리 설정
        cw.w61()  # 원격에서 이벤트 로그 파일 접근 차단
        cw.w62()  # 백신 프로그램 설치
        cw.w63()  # SAM 파일 접근 통제 설정
        cw.w64()  # 화면보호기설정
        cw.w65()  # 로그온하지 않고 시스템 종료 허용
        cw.w66()  # 원격 시스템에서 강제로 시스템 종료
        cw.w67()  # 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료
        cw.w68()  # SAM 계정과 공유의 익명 열거 허용 안 함
        cw.w69()  # Autologon 기능 제어
        cw.w70()  # 이동식 미디어 포맷 및 꺼내기 허용
        cw.w71()  # 디스크볼륨 암호화 설정 / 측정 불가
        cw.w72()  # Dos공격 방어 레지스트리 설정 / 확인 완료 2017-11-14(화)
        cw.w73()  # 사용자가 프린터 드라이버를 설치할 수 없게 함 / 확인 완료 2017-11-14(화)
        cw.w74()  # 세션 연결을 중단하기 전에 필요한 유휴시간 / 확인 완료 2017-11-14(화)
        cw.w75()  # 경고 메시지 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w76()  # 사용자 별 홈 디렉터리 권한 설정 / 확인 완료 2017-11-14(화)
        cw.w77()  # LAN Manager 인증 수준 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w78()  # 보안 채널 데이터 디지털 암호화 또는 서명 / 확인 완료 2017-11-14(화)
        cw.w79()  # 파일 및 디렉터리 보호 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w80()  # 컴퓨터 계정 암호 최대 사용 기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w81()  # 시작프로그램 목록 분석  /  확인 완료 2017-11-13(월)
        cw.w82()   # Windows 인증 모드 사용
        self.excel(cw)
        self.htmlFile(cw)

    def run_Iis(self):
        cw = Check_Vulnerability_Exposures()
        cw.w01()  # Administrator 계정 이름 바꾸기 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w02()  # Guest 계정 상태 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w03()  # 불필요한 계정 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w04()  # 계정 잠금 임계값 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w05()  # 해독 가능한 암호화를 사용하여 암호 저장 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w06()  # 관리자 그룹에 최소한의 사용자 포함 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w07()  # Everyone 사용 권한을 익명 사용자에게 적용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w08()  # 계정 잠금 기간 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w09()  # 패스워드 복잡성 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w10()  # 패스워드 최소 암호 길이 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w11()  # 패스워드 최대 사용 기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w12()  # 패스워드최소사용기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w13()  # 마지막 사용자 이름 표시 안함 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w14()  # 로컬 로그온 허용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w15()  # 익명 SID/이름 변환 허용 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w16()  # 최근 암호 기억 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w17()  # 콘솔 로그온 시 로컬 계정에서 빈 암호 사용 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w18()  # 원격터미널 접속 가능한 사용자 그룹 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w19()  # 공유 권한 및 사용자 그룹 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w20()  # 하드디스크 기본 공유 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w21()  # 불필요한 서비스 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w22()  # IIS 서비스 구동 점검 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w23()  # IIS 디렉토리 리스팅 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w24()  # IIS CGI 실행 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w25()  # IIS 상위 디렉토리 접근 금지 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w26()  # IIS 불필요한 파일 제거 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w27()  # IIS 웹 프로세스 권한 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w28()  # IIS 링크 사용금지 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w29()  # IIS 파일 업로드 및 다운로드 제한 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w30()  # IIS DB 연결 취약점 점검 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w31()  # IIS 가상 디렉토리 삭제
        cw.w32()  # IIS 데이터 파일 ACL 적용
        cw.w33()  # IIS 미사용 스크립트 매핑 제거
        cw.w34()  # IIS Exec 명령어 쉘 호출 진단
        cw.w35()  # IIS WebDAV 비활성화
        cw.w36()  # NetBIOS 바인딩 서비스 구동 점검
        if self.ftp_comp is True:
            cw.w37()  # FTP 서비스 구동 점검
            cw.w38()  # FTP 디렉토리 접근권한 설정
            cw.w39()  # Anonymouse FTP 금지
            cw.w40()  # FTP 접근 제어 설정
        else:
            for ftp_numb in range(37,41):
                self.ftp_not_Remainder(ftp_numb,cw)
        cw.w41()  # DNS Zone Transfer 설정
        self.iis_not_Remainder(42, cw)  # IIS 42
        cw.w43()  # 최신 서비스팩 적용
        cw.w44()  # 터미널 서비스 암호화 수준 설정
        self.iis_not_Remainder(45, cw) # IIS 45
        self.iis_not_Remainder(46, cw) # IIS 46
        cw.w47()  # SNMP 서비스 커뮤니티스트링의 복잡성 설정
        cw.w48()  # SNMP Access control 설정
        cw.w49()  # DNS 서비스 구동 점검
        self.iis_not_Remainder(50, cw) # IIS 50
        cw.w51()  # Telnet 보안 설정
        cw.w52()  # 불필요한 ODBC/OLE-DB 데이터 소스와 드라이브 제거
        cw.w53()  # 원격터미널 접속 타임아웃 설정
        cw.w54()  # 예약된 작업에 의심스러운 명령이 등록되어 있는지 점검
        cw.w55()  # 최신 HOT FIX 적용
        cw.w56()  # 백신 프로그램 업데이트
        cw.w57()  # 정책에 따른 시스템 로깅 설정
        cw.w58()  # 로그의 정기적 검토 및 보고
        cw.w59()  # 원격으로 액세스할 수 있는 레지스트리 경로
        cw.w60()  # 이벤트 로그 관리 설정
        cw.w61()  # 원격에서 이벤트 로그 파일 접근 차단
        cw.w62()  # 백신 프로그램 설치
        cw.w63()  # SAM 파일 접근 통제 설정
        cw.w64()  # 화면보호기설정
        cw.w65()  # 로그온하지 않고 시스템 종료 허용
        cw.w66()  # 원격 시스템에서 강제로 시스템 종료
        cw.w67()  # 보안 감사를 로그할 수 없는 경우 즉시 시스템 종료
        cw.w68()  # SAM 계정과 공유의 익명 열거 허용 안 함
        cw.w69()  # Autologon 기능 제어
        cw.w70()  # 이동식 미디어 포맷 및 꺼내기 허용
        cw.w71()  # 디스크볼륨 암호화 설정 / 측정 불가
        cw.w72()  # Dos공격 방어 레지스트리 설정 / 확인 완료 2017-11-14(화)
        cw.w73()  # 사용자가 프린터 드라이버를 설치할 수 없게 함 / 확인 완료 2017-11-14(화)
        cw.w74()  # 세션 연결을 중단하기 전에 필요한 유휴시간 / 확인 완료 2017-11-14(화)
        cw.w75()  # 경고 메시지 설정 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w76()  # 사용자 별 홈 디렉터리 권한 설정 / 확인 완료 2017-11-14(화)
        cw.w77()  # LAN Manager 인증 수준 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w78()  # 보안 채널 데이터 디지털 암호화 또는 서명 / 확인 완료 2017-11-14(화)
        cw.w79()  # 파일 및 디렉터리 보호 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w80()  # 컴퓨터 계정 암호 최대 사용 기간 / AttributeError 적용 완료 2017-11-17 (금)
        cw.w81()  # 시작프로그램 목록 분석  /  확인 완료 2017-11-13(월)
        cw.w82()  # Windows 인증 모드 사용
        self.excel(cw)
        self.htmlFile(cw)

    def iis_not_Remainder(self,code_number,cw):
        cw.status_code['W-{}'.format(code_number)] = "양호"
        result = "[W-{}] IIS Server를 사용하고 있지 않습니다.".format(code_number)
        cw.result_list['W-{}'.format(code_number)] = result

    def ftp_not_Remainder(self,code_number,cw):
        cw.status_code['W-{}'.format(code_number)] = "양호"
        result = "[W-{}] FTP Service를 사용하고 있지 않습니다.".format(code_number)
        cw.result_list['W-{}'.format(code_number)] = result

    def service_status(self,service_name):
        flag = False
        try:
            win32serviceutil.QueryServiceStatus(service_name)
            flag = True
        except pywintypes.error:
            pass
        return flag

    def htmlFile(self,cw):
        with open("sample/item.txt", "r", encoding='cp949') as f:
            items = f.readlines()
        with codecs.open("sample/window_result.html", "r", encoding='utf-8') as f:
            total = f.read()
        result = ""
        for cnt in range(len(items)):
            num = cnt + 1
            eitem = items[cnt].replace("\n", "")
            result_str = cw.result_list.get('W-'+'{}'.format(num).zfill(2))
            code = cw.status_code.get('W-'+'{}'.format(num).zfill(2))
            if code =="취약":
                color = "style='background-color:red';"
            elif code == "양호":
                color = "style='background-color:aqua';"
            else:
                color = "style='background-color:#FFFFFF';"
            tab = "\t" * 4
            tab2 = "\t" * 5
            if cnt == 0:
                result = result + tab + "<tr id='service1'>\n" + tab2 + "<td rowspan='18' class='kind1'>계정 관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            elif cnt == 18:
                result = result + tab + "<tr id='service2'>\n" + tab2 + "<td rowspan='36' class='kind2'>서비스\n관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            elif cnt == 54:
                result = result + tab + "<tr id='service3'>\n" + tab2 + "<td rowspan='3' class='kind3'>패치 관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            elif cnt == 57:
                result = result + tab + "<tr id='service4'>\n" + tab2 + "<td rowspan='4' class='kind4'>로그 관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            elif cnt == 61:
                result = result + tab + "<tr id='service5'>\n" + tab2 + "<td rowspan='20' class='kind5'>보안 관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            elif cnt == 81:
                result = result + tab + "<tr id='service6'>\n" + tab2 + "<td rowspan='1' class='kind6'>DB 관리</td>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(
                    num).zfill(2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"
            else:
                result = result + tab + "<tr>\n" + tab2 + "<td class='wcode'>W-" + "{}".format(num).zfill(
                    2) + "</td>\n" + tab2 + "<td class='item'>{}</td>\n".format(
                    eitem) + tab2 + "<td class='result_text'>{}</td>\n".format(result_str) + tab2 + "<td class='result_status'{0}>{1}</td>\n".format(color,code) + tab + "</tr>\n"

        a = total.replace("replace", result)
        with open("sample/window_Confirm.html", "w",encoding='utf-8') as f:
            f.write(a)
        print("HTML File 생성!")

    def excel(self,cw):
        ed = openpyxl.load_workbook("sample/vulner_report.xlsx")
        sheet = ed.get_sheet_by_name('Windows 점검')
        for x in range(len(cw.result_list)):
            status = cw.status_code.get('W-'+'{}'.format(x+1).zfill(2))
            result = cw.result_list.get('W-'+'{}'.format(x+1).zfill(2))
            sheet["E{}".format(x + 4)].alignment = Alignment(wrapText=True) # 줄바꿈 사용
            sheet["E{}".format(x+4)] = result # 한 셀의 값은 리스트의 값으로 지정
            sheet["F{}".format(x+4)].alignment = Alignment(horizontal='center',vertical='center')
            sheet["F{}".format(x+4)] = status # 상태코드 삽입
            if status == "취약":
                sheet["F{}".format(x+4)].fill = PatternFill(patternType='solid',fgColor=Color('FF0000')) # 배경색
                sheet["F{}".format(x + 4)].font = Font(bold=True)
            elif status == "양호":
                sheet["F{}".format(x + 4)].fill = PatternFill(patternType='solid', fgColor=Color('00B0F0'))  # 배경색
            else:
                # sheet["F{}".format(x + 4)].fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))  # 배경색
                sheet["F{}".format(x + 4)].font = Font(bold=True)
        try:

            ed.save("Check-Win_Result! [{}].xlsx".format(self.now))
        except PermissionError:
            print("Excel 창이 실행중입니다.")
        self.common_func("chcp 949")
        print("엑셀 파일 생성!")


if __name__ == '__main__':
    start = time.time()
    # cve = Check_Vulnerability_Exposures()
    cw = Check_Win()
    cw.start()
    # cw.excel()
    # cve.play_cve()  # 진단
    # cve.excel()
    os.remove(os.getcwd() + "\Local_Security_Policy.txt")
    print(int(time.time() - start), "초")